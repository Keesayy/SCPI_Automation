import pyvisa
import openpyxl
import time

### Constants
PRECISION = 3
NB_ESE_BITS = 60
NB_QUES_BITS = 952
DEBUG = False

def PRINT(*args, **kwargs):
	if(DEBUG):
		return __builtins__.print(*args, **kwargs)

def Float_precision_str(n : int) -> str:
	s : str = '#,##0.'
	for i in range(n):
		s += '0'
	return s

def Excel_name(name : str, precision : int, freq_start : float, freq_stop : float, nb_points : int, dwel : float, amplitude : float) -> str:
	s : string = name
	s = s + '-' + ('{:.%df}' % precision).format(freq_start)
	s = s + '-' + ('{:.%df}' % precision).format(freq_stop)
	s = s + '-' + str(nb_points)
	s = s + '-' + ('{:.%df}' % precision).format(dwel)
	s = s + '-' + ('{:.%df}' % precision).format(amplitude)
	s = s + '.xlsx'
	return s;

def Gpid_devices_open():
	rm = pyvisa.ResourceManager()
	print(rm.list_resources(), '\n')

	#Open gpid devices
	power_meter = rm.open_resource('GPIB0::13::INSTR')
	signal_source = rm.open_resource('GPIB0::19::INSTR')

	power_meter.write('SYST:LANG SCPI')
	# signal_source.write('SYST:LANG SCPI')
	signal_source.write('/SCPI')
	time.sleep(0.2)

	print(power_meter.query('*IDN?'), end = "")
	print(signal_source.query('*IDN?'), end = "")
	return rm ,power_meter, signal_source

def Signal_source_init(signal_source) -> None:
	### Signal Source Init
	# signal_source.write('*RST')
	signal_source.write('*CLS')
	signal_source.write('FREQ:MODE LIST')
	signal_source.write('POW:MODE LIST')
	signal_source.write('LIST:TYPE STEP')

	signal_source.write('POW:ATT:AUTO ON')
	signal_source.write('POW:ATT 0 DB')
	signal_source.write('POW:ALC:LEV 0 DB')

def Power_meter_init(power_meter) -> None:
	### Power Meter Init
	NB_DIGIT = 3
	power_meter.write('*CLS')
	power_meter.write('*ESE 1') 
	power_meter.write('UNIT:POW dBm')

	PRINT('ESR : ', power_meter.write('*ESR?'))
	# power_meter.write('SENSE:AVER:COUN:AUTO ON')
	# PRINT('ESR : ', power_meter.query('*ESR?'))

	power_meter.write('DISP:RES %d' % NB_DIGIT)

def Show_parameters(freq_start : float, freq_stop : float, nb_points : int, dwel : float, amplitude : float) -> None:
	print('\nSTARTING ACQUISITION WITH PARAMETERS :')
	print('fstart :', ('{:.%df}' % PRECISION).format(freq_start), 'Ghz')
	print('fstop  :', ('{:.%df}' % PRECISION).format(freq_stop), 'Ghz')
	print('points :' + str(nb_points))
	print('dwel   :', ('{:.%df}' % PRECISION).format(dwel), 'ms')
	print('amp    :', ('{:.%df}' % PRECISION).format(amplitude), 'dBm')
	print('\n')

def Sweep_freq(power_meter, signal_source, freq_start : float, freq_stop : float, nb_points : int, dwel : float, amplitude : float):

	# signal_source.write('FREQ:STAR %f GHz' % freq_start)
	# print(signal_source.query('FREQ:STAR?'))

	# signal_source.write('FREQ:STOP %f GHz' % freq_stop)
	# print(signal_source.query('FREQ:STOP?'))

	# signal_source.write('SWE:POIN %d' % nb_points)
	# signal_source.write('SWE:DWEL %f MS' % dwel)

	# signal_source.write('POW:AMPL %f dBm' % amplitude)

	# signal_source.write('LIST:TRIG:SOUR BUS')
	# signal_source.write('OUTP:STAT ON')
	# signal_source.write('INIT:CONT ON')

	signal_source.write('OUTP ON')
	print(signal_source.query('*OPC?'))
	signal_source.write('POW %f dBm' % amplitude)

	# Load the Excel workbook
	excel = openpyxl.load_workbook(filename = 'E8257D-67 2024.xlsx')
	PRINT(excel.sheetnames)
	sheet = excel[excel.sheetnames[1]]

	freq = freq_start
	precision_string = Float_precision_str(PRECISION)

	for i in range(nb_points):
		signal_source.write('FREQ:CW %f GHz' % freq)

		power_meter.write('*CLS')
		power_meter.write('FREQ ' + str(freq) + ' GHz')

		#Measure level
		power_meter.write('TRIG:DEL:AUTO ON')
		power_meter.write('INIT:CONT OFF') 
		power_meter.write('TRIG:SOUR IMM') 
		power_meter.write('INIT') 

		power_meter.write('*OPC')

		start = time.time()
		stb_polling(power_meter, signal_source, timeout = 20, sleepTime = 0.15)
		end = time.time()

		#Clear ESE
		power_meter.query('*ESR?') 

		#Read Level
		level = power_meter.query('FETCH?')

		print(str(i) + ':', ('{:.%df}' % PRECISION).format(float(freq)), 'GHz | ', ('{:.%df}' % PRECISION).format(float(level)) , 'dBm | ' , ('{:.%df}' % PRECISION).format(end - start), 's')

		#Excel array
		sheet['B' + str(i+3)] = freq
		sheet['C' + str(i+3)] = float(('{:.6f}').format(float(level)))
		sheet['B' + str(i+3)].number_format = precision_string
		sheet['C' + str(i+3)].number_format = precision_string

		# signal_source.write('*TRG')

		freq += (freq_stop - freq_start) / (nb_points - 1)

	# signal_source.write('INIT:CONT OFF')
	# signal_source.write('OUTP:STAT OFF')
	signal_source.write('OUTP OFF')

	# Save the workbook
	new_workbook_name = Excel_name('Giga-2420B', 0, freq_start, freq_stop, nb_points, dwel, amplitude)	
	excel.save(new_workbook_name)
	return excel, new_workbook_name

def CLOSE_ALL(signal_source, power_meter, excel, rm) -> None:
	signal_source.close()
	power_meter.close()
	excel.close()
	rm.close()

def stb_polling(instrument, instrument_bis, condition = 32, timeout = 1.0, sleepTime = 0.3):
	PRINT('Polling started')
	end_time = time.time() + timeout # compute the maximal end time
	status = False
	error = False
	stb = instrument.read_stb()
	status = (stb & condition) == condition # first condition check, no need to wait if condition already true
	error = (stb & 1) == 1 # check error, no need to wait if already an error is available

	while not status and time.time() < end_time and not error:
		PRINT('STB : ', stb)
		time.sleep(sleepTime)
		stb = instrument.read_stb()
		status = (stb & condition) == condition #check conditon
		error = (stb & 4) == 4 # check bit 4: Error Message Available
	if status:
		PRINT('Polling finished because STB satisfied condition. STB = ', condition)
	elif error:
		print('Polling finished because Error Occured')
		print(instrument.query('SYST:ERR?'))
		print(instrument_bis.query('SYST:ERR?'))
	else:
		print('Polling finished because timeout of', timeout, 'seconds reached')
	return status	

def main() -> int:
	rm, power_meter, signal_source = Gpid_devices_open()

	Signal_source_init(signal_source)
	Power_meter_init(power_meter)

	# Step_Sweep
	freq_start : float = 2     	#GHZ
	freq_stop  : float = 20		#GHZ
	nb_points  : int   = 900			
	dwel       : float = 1.00	#MS
	amplitude  : float = 3		#DBM 

	start_tot = time.time()

	print('\nSTART ACQUISITION SWEEP FREQ')
	Show_parameters(freq_start, freq_stop, nb_points, dwel, amplitude)
	
	excel, excel_name = Sweep_freq(power_meter, signal_source, freq_start, freq_stop, nb_points, dwel, amplitude)

	end_tot = time.time()
	print('\nTOTAL ACQUISITION TIME : ', ('{:.%df}' % PRECISION).format(end_tot - start_tot), 's')
	print('Saved in : ', excel_name)
	CLOSE_ALL(signal_source, power_meter, excel, rm)
	return 0

if __name__=="__main__":
    main()







